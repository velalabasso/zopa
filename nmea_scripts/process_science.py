#!/usr/bin/env python3
"""
process_science.py  —  Vela Lab Science Export
===============================================
Intégration dans nmea_process.py :
  1. En tête (avec les imports) :
         from process_science import run_science_export

  2. En fin de boucle, après generate_pdf(...) :
         run_science_export(df, df_events, meta, log_dir, base)

Sortie : <log_dir>/<base>_science.xlsx

Lignes produites :
  - 1 ligne HYP par paire STATION HYP ON → OFF  (avec turbidity et secchi si présents)
  - 1 ligne BIO par paire FILTRATION ON → OFF    (avec planctospace, size, volume)
  - 1 ligne BUCKET par event BUCKET dans la fenêtre de la station BIO active
"""

import os
import re
import math
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# TIMESTAMP  — tz-safe
# ─────────────────────────────────────────────────────────────────────────────

def _to_ns(ts) -> int:
    try:
        t = pd.Timestamp(ts)
        if t.tzinfo is not None:
            t = t.tz_convert("UTC").tz_localize(None)
        return t.value
    except Exception:
        return 0


def _normalize_events(df_events):
    if df_events.empty or "timestamp" not in df_events.columns:
        return df_events
    df = df_events.copy()
    df["timestamp"] = df["timestamp"].apply(
        lambda x: pd.Timestamp(_to_ns(x)) if pd.notna(x) else pd.NaT
    )
    return df


def _series_to_ns(s: pd.Series) -> np.ndarray:
    s2 = pd.to_datetime(s)
    if s2.dt.tz is not None:
        s2 = s2.dt.tz_convert("UTC").dt.tz_localize(None)
    return s2.values.astype("datetime64[ns]").astype("int64")


# ─────────────────────────────────────────────────────────────────────────────
# FORMATAGE  (virgule comme séparateur décimal)
# ─────────────────────────────────────────────────────────────────────────────

def _fmt(value, decimals=6) -> str:
    if value is None or value == "":
        return ""
    try:
        v = float(value)
        if math.isnan(v):
            return ""
        return f"{v:.{decimals}f}".replace(".", ",")
    except (TypeError, ValueError):
        return str(value)

def _fmt2(v): return _fmt(v, 2)
def _fmt4(v): return _fmt(v, 4)
def _fmt6(v): return _fmt(v, 6)
def _fmt8(v): return _fmt(v, 8)


# ─────────────────────────────────────────────────────────────────────────────
# GÉOGRAPHIE
# ─────────────────────────────────────────────────────────────────────────────

def _decimal_to_dms(decimal: float, is_lon: bool = False) -> str:
    if decimal is None or (isinstance(decimal, float) and math.isnan(decimal)):
        return ""
    direction = ("E" if decimal >= 0 else "W") if is_lon else ("N" if decimal >= 0 else "S")
    a = abs(decimal)
    deg = int(a)
    minutes = f"{(a - deg) * 60.0:.4f}".replace(".", ",")
    dw = 3 if is_lon else 2
    return f"{deg:0{dw}d}°{minutes}'{direction}"


def _haversine_km(lat1, lon1, lat2, lon2) -> float:
    R = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    a = (math.sin(math.radians(lat2 - lat1) / 2) ** 2
         + math.cos(phi1) * math.cos(phi2)
         * math.sin(math.radians(lon2 - lon1) / 2) ** 2)
    return 2 * R * math.atan2(math.sqrt(a), math.sqrt(1 - a))


# ─────────────────────────────────────────────────────────────────────────────
# INTERPOLATION NMEA
# ─────────────────────────────────────────────────────────────────────────────

def _build_nmea_arrays(df_nmea):
    if df_nmea.empty or "datetime" not in df_nmea.columns:
        return np.array([]), np.array([]), np.array([]), np.array([]), False
    df_t = df_nmea.dropna(subset=["datetime", "lat_raw", "lon_raw"]).copy()
    df_t = df_t.sort_values("datetime").reset_index(drop=True)
    if df_t.empty:
        return np.array([]), np.array([]), np.array([]), np.array([]), False
    ts_ns = _series_to_ns(df_t["datetime"])
    lat   = df_t["lat_raw"].values.astype(float)
    lon   = df_t["lon_raw"].values.astype(float)
    sog   = (df_t["SOG_RMC"].values.astype(float)
             if "SOG_RMC" in df_t.columns else np.full(len(df_t), np.nan))
    return ts_ns, lat, lon, sog, True


def _interp(ts_ns_arr, lat_arr, lon_arr, sog_arr, t_ns):
    if len(ts_ns_arr) == 0:
        return np.nan, np.nan, np.nan
    idx = np.searchsorted(ts_ns_arr, t_ns, side="left")
    if idx == 0:
        return float(lat_arr[0]), float(lon_arr[0]), float(sog_arr[0])
    if idx >= len(ts_ns_arr):
        return float(lat_arr[-1]), float(lon_arr[-1]), float(sog_arr[-1])
    dt = ts_ns_arr[idx] - ts_ns_arr[idx - 1]
    if dt == 0:
        return float(lat_arr[idx]), float(lon_arr[idx]), float(sog_arr[idx])
    a = (t_ns - ts_ns_arr[idx - 1]) / dt
    lat = float(lat_arr[idx-1]) + a * (float(lat_arr[idx]) - float(lat_arr[idx-1]))
    lon = float(lon_arr[idx-1]) + a * (float(lon_arr[idx]) - float(lon_arr[idx-1]))
    sog = float(sog_arr[idx-1]) + a * (float(sog_arr[idx]) - float(sog_arr[idx-1]))
    return lat, lon, sog


def _mean_sog(ts_ns_arr, sog_arr, t0, t1) -> float:
    mask = (ts_ns_arr >= t0) & (ts_ns_arr <= t1)
    vals = sog_arr[mask]
    vals = vals[~np.isnan(vals)]
    return float(np.mean(vals)) if len(vals) > 0 else np.nan


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES BIO
# ─────────────────────────────────────────────────────────────────────────────

_MOUTH_DIAM_LAMPREY = "6"   # cm — Lamprey filtrations
_DEPTH_BIO          = "1"   # m  — toutes stations BIO (lamprey + bucket)

_SIZE_RANGE = {
    "micro": ("50",  "200"),
    "nano":  ("5",   "50"),
    "pico":  ("0,3", "5"),
}

def _size_min_max(size: str):
    """Retourne (size_min, size_max) pour micro/nano/pico, sinon ("", "")."""
    return _SIZE_RANGE.get(size.lower().strip(), ("", ""))


# ─────────────────────────────────────────────────────────────────────────────
# PARSE DES DÉTAILS D'EVENTS
# ─────────────────────────────────────────────────────────────────────────────

def _parse_filtration_off(detail: str) -> dict:
    """
    Parse : ' OFF | STATION Vela_Lab_bio_st01 | SIZE MICRO | VOLUME 500,0 mL
              | SATURATION NO | PLANCTOSPACE st06-03'
    Retourne dict avec keys: size, volume_ml, saturation, planctospace
    """
    result = {"size": "", "volume_ml": "", "saturation": "", "planctospace": ""}
    for part in detail.split("|"):
        part = part.strip()
        if part.upper().startswith("SIZE"):
            result["size"] = part.split(None, 1)[-1].strip().lower()
        elif part.upper().startswith("VOLUME"):
            m = re.search(r"([\d,\.]+)", part)
            if m:
                result["volume_ml"] = m.group(1).replace(".", ",")
        elif part.upper().startswith("SATURATION"):
            result["saturation"] = part.split(None, 1)[-1].strip()
        elif part.upper().startswith("PLANCTOSPACE"):
            ps = part.split(None, 1)[-1].strip()
            # Garder uniquement XX de st06-XX
            m = re.search(r"(?:st\d+-)?(\w+)$", ps, re.IGNORECASE)
            result["planctospace"] = m.group(1) if m else ps
    return result


def _parse_filtration_on(detail: str) -> dict:
    """
    Parse : ' ON | STATION Vela_Lab_bio_st01 | SIZE MICRO'
    """
    result = {"size": ""}
    for part in detail.split("|"):
        part = part.strip()
        if part.upper().startswith("SIZE"):
            result["size"] = part.split(None, 1)[-1].strip().lower()
    return result


def _parse_turbidity(detail: str) -> dict:
    """
    Parse : ' | STATION Vela_Lab_hyp_st001 | T1=0,1234 | T2=0,5678 | T3=0,9012'
    Retourne dict t1, t2, t3
    """
    result = {"t1": "", "t2": "", "t3": ""}
    for part in detail.split("|"):
        part = part.strip()
        ul = part.upper()
        if ul.startswith("T1="):
            result["t1"] = part[3:].replace(".", ",")
        elif ul.startswith("T2="):
            result["t2"] = part[3:].replace(".", ",")
        elif ul.startswith("T3="):
            result["t3"] = part[3:].replace(".", ",")
    return result


def _parse_secchi(detail: str) -> str:
    """
    Parse : ' | STATION Vela_Lab_hyp_st001 | DEPTH 12,50 m'
    Retourne la valeur de profondeur.
    """
    for part in detail.split("|"):
        part = part.strip()
        if part.upper().startswith("DEPTH"):
            m = re.search(r"([\d,\.]+)", part)
            if m:
                return m.group(1).replace(".", ",")
    return ""


def _station_id_from_detail(detail: str, prefix: str) -> str:
    """
    Extrait 'Vela_Lab_bio_st01' ou 'Vela_Lab_hyp_st001' d'un detail d'event.
    prefix = 'bio' ou 'hyp'
    """
    m = re.search(rf"Vela_Lab_{prefix}_st(\S+)", detail, re.IGNORECASE)
    if m:
        return f"Vela_Lab_{prefix}_st{m.group(1)}"
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# COLONNES
# ─────────────────────────────────────────────────────────────────────────────

COLUMNS = [
    ("station_id",      "Stations id",           "Vela Lab station",               "auto"),
    ("station_type",    "Stations id",           "Vela Lab station type",          "auto"),
    ("cloud_coverage",  "Stations id",           "cloud_coverage (%)",             "manual"),
    ("date_start",      "Start",                 "date_start (YYYY-MM-DD)",        "auto"),
    ("time_start",      "Start",                 "time_start (UTC HH:MM:SS)",      "auto"),
    ("lat_start_dms",   "Start",                 "lat_start (°N)",                 "auto"),
    ("lon_start_dms",   "Start",                 "lon_start (°E)",                 "auto"),
    ("lat_start_dec",   "Start",                 "lat_start (decimal °N)",         "auto"),
    ("lon_start_dec",   "Start",                 "lon_start (decimal °E)",         "auto"),
    ("date_end",        "End",                   "date_end (YYYY-MM-DD)",          "auto"),
    ("time_end",        "End",                   "time_end (UTC HH:MM:SS)",        "auto"),
    ("lat_end_dms",     "End",                   "lat_end (°N)",                   "auto"),
    ("lon_end_dms",     "End",                   "lon_end (°E)",                   "auto"),
    ("lat_end_dec",     "End",                   "lat_end (decimal °N)",           "auto"),
    ("lon_end_dec",     "End",                   "lon_end (decimal °E)",           "auto"),
    ("device",          "End",                   "device",                         "auto"),
    ("mouth_diam",      "End",                   "mouth_Ø (cm)",                   "auto"),
    ("depth",           "Depth",                 "depth (m)",                      "auto"),
    ("size_min",        "Depth",                 "size_min (um)",                  "auto"),
    ("size_max",        "Comments",              "size_max (um)",                  "auto"),
    ("comment",         "Comments",              "comment about the sampling",     "manual"),
    ("duration_min",    "Duration & Distance",   "duration (min)",                 "auto"),
    ("dist_km",         "Duration & Distance",   "distance (km) Haversine",        "auto"),
    ("sog_haversine",   "Duration & Distance",   "SOG_moy (kts)",                  "auto"),
    ("sog_nmea",        "Duration & Distance",   "SOG_moy NMEA (kts)",             "auto"),
    ("surface_mouth",   "Volume",                "surface mouth (m²)",             "formula"),
    ("vol_net",         "Volume",                "vol_net theoric (m³)",           "formula"),
    ("vol_net_conc",    "Volume",                "vol_net_conc (ml)",              "manual"),
    ("planktospace_st", "PlanktoSpace station",  "PlanktoSpace station",           "auto"),
    ("barcode",         "PlanktoSpace station",  "barcode",                        "manual"),
    ("vol_lamprey",     "Lamprey analysis",      "size (micro/nano/pico)",         "auto"),
    ("time_filt_start", "Lamprey analysis",      "time_filtration_start (UTC)",    "auto"),
    ("time_filt_end",   "Lamprey analysis",      "time_filtration_end (UTC)",      "auto"),
    ("saturation",      "Lamprey analysis",      "saturation (Y/N)",               "auto"),
    ("vol_lamprey_fil", "Lamprey analysis",      "vol_lamprey_filtrate (ml)",      "auto"),
    ("vol_curiosity",   "QCuriosity analysis",   "vol_curiosity (ml)",             "manual"),
    ("conc_dil",        "QCuriosity analysis",   "conc_or_dilution",               "manual"),
    ("nb_images",       "QCuriosity analysis",   "nb_images",                      "manual"),
    ("vol_pscope",      "Planktoscope analysis", "vol_pscope (ml)",                "manual"),
    ("acq",             "Planktoscope analysis", "acq",                            "manual"),
    ("seg",             "Planktoscope analysis", "seg",                            "manual"),
    ("facteur_pscope",  "Planktoscope analysis", "facteur_pscope",                 "manual"),
    ("vol_pscope_eff",  "Planktoscope analysis", "vol_pscope_effectif (ml)",       "manual"),
    ("vol_imaged",      "Planktoscope analysis", "vol_imaged (ml)",                "manual"),
    ("acq_imaged_vol",  "Planktoscope analysis", "acq_imaged_volume (ml)",         "manual"),
    ("conc_filet",      "Planktoscope analysis", "conc_filet (m³/organisme)",      "manual"),
    ("time_turbi",      "Turbidometre",          "time_turbi (UTC HH:MM:SS)",      "auto"),
    ("turbi_1",         "Turbidometre",          "turbi_1",                        "auto"),
    ("turbi_2",         "Turbidometre",          "turbi_2",                        "auto"),
    ("turbi_3",         "Turbidometre",          "turbi_3",                        "auto"),
    ("secchi",          "Secchi Disk",           "Secchi Disk",                    "auto"),
]


# ─────────────────────────────────────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────────────────────────────────────

_NAVY       = "1A3A5C"
_WHITE      = "FFFFFF"
_LGRAY      = "F0F4FA"
_LBLUE      = "D6EAF8"
_GREEN_PALE = "EAF6EE"
_YELLOW     = "FFF3CD"
_ORANGE     = "FFE0B2"

def _hfill(h):      return PatternFill("solid", fgColor=h)
def _font(bold=False, color="000000", size=9):
    return Font(bold=bold, color=color, size=size, name="Arial")
def _border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)
def _center():      return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left():        return Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ─────────────────────────────────────────────────────────────────────────────
# ÉCRITURE XLSX
# ─────────────────────────────────────────────────────────────────────────────

def _write_xlsx(rows: list, out_path: str, leg_label: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Science"
    n_cols = len(COLUMNS)

    # Ligne 1 — groupes fusionnés
    groups, prev_g, start_c = [], None, 1
    for i, (_, g, _, _) in enumerate(COLUMNS):
        if g != prev_g:
            if prev_g is not None:
                groups.append((prev_g, start_c, i))
            prev_g, start_c = g, i + 1
    groups.append((prev_g, start_c, n_cols))

    for g_label, c_start, c_end in groups:
        cell = ws.cell(row=1, column=c_start, value=g_label)
        cell.fill = _hfill(_NAVY); cell.font = _font(bold=True, color=_WHITE, size=9)
        cell.alignment = _center(); cell.border = _border()
        if c_end > c_start:
            ws.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)

    # Ligne 2 — noms de colonnes
    for i, (_, _, col_label, origin) in enumerate(COLUMNS):
        bg = _LBLUE if origin == "auto" else (_YELLOW if origin == "formula" else _LGRAY)
        cell = ws.cell(row=2, column=i+1, value=col_label)
        cell.fill = _hfill(bg); cell.font = _font(bold=True, size=8)
        cell.alignment = _center(); cell.border = _border()

    # Indices pour formules
    mouth_ltr = get_column_letter(next(i+1 for i,(k,*_) in enumerate(COLUMNS) if k=="mouth_diam"))
    dist_ltr  = get_column_letter(next(i+1 for i,(k,*_) in enumerate(COLUMNS) if k=="dist_km"))
    surf_ltr  = get_column_letter(next(i+1 for i,(k,*_) in enumerate(COLUMNS) if k=="surface_mouth"))

    # Données
    for r_idx, session in enumerate(rows):
        row    = 3 + r_idx
        row_bg = session.get("_row_bg", None)

        for col_idx, (key, _, _, origin) in enumerate(COLUMNS):
            col = col_idx + 1
            if key == "surface_mouth":
                val = f'=IF({mouth_ltr}{row}<>"",PI()*(({mouth_ltr}{row}/100)/2)^2,"")'
            elif key == "vol_net":
                val = f'=IF(AND({surf_ltr}{row}<>"",{dist_ltr}{row}<>""),{surf_ltr}{row}*{dist_ltr}{row},"")'
            else:
                val = session.get(key, "")
                if val is None or val is np.nan:
                    val = ""

            cell = ws.cell(row=row, column=col, value=val)
            cell.border = _border(); cell.alignment = _left(); cell.font = _font(size=9)
            if origin == "formula":
                cell.fill = _hfill(_YELLOW)
            elif origin == "auto" and val not in (None, "", np.nan):
                cell.fill = _hfill(row_bg if row_bg else _GREEN_PALE)
            elif row_bg:
                cell.fill = _hfill(row_bg)

    # Légende
    lr = 3 + len(rows) + 2
    for i, (color, label) in enumerate([
        (_GREEN_PALE, "Rempli automatiquement depuis NMEA / events"),
        (_YELLOW,     "Calculé par formule Excel (saisir le Ø filet pour activer)"),
        (_ORANGE,     "Ligne Bucket (position = moment du bucket)"),
        (_WHITE,      "À remplir manuellement après la mesure"),
    ]):
        c = ws.cell(row=lr+i, column=1, value=label)
        c.fill = _hfill(color); c.font = _font(size=8); c.border = _border()

    ws.cell(row=lr+6, column=1,
            value=f"Généré par process_science.py — {leg_label} — "
                  f"{len(rows)} ligne(s). "
                  "Vérifier puis copier-coller dans Bio_sampling_Vela_Lab (Google Drive)."
            ).font = _font(size=8, color="666666")

    # Largeurs
    widths = {
        "station_id":22,"station_type":12,"cloud_coverage":10,
        "date_start":14,"time_start":14,
        "lat_start_dms":22,"lon_start_dms":22,
        "lat_start_dec":16,"lon_start_dec":16,
        "date_end":14,"time_end":14,
        "lat_end_dms":22,"lon_end_dms":22,
        "lat_end_dec":16,"lon_end_dec":16,
        "device":16,"mouth_diam":10,"depth":8,
        "size_min":9,"size_max":9,"comment":28,
        "duration_min":11,"dist_km":16,
        "sog_haversine":13,"sog_nmea":13,
        "surface_mouth":14,"vol_net":18,
        "planktospace_st":18,
        "vol_lamprey":14,"time_filt_start":16,"time_filt_end":16,
        "saturation":10,"vol_lamprey_fil":16,
        "time_turbi":16,"turbi_1":10,"turbi_2":10,"turbi_3":10,"secchi":10,
    }
    for i, (key,*_) in enumerate(COLUMNS):
        ws.column_dimensions[get_column_letter(i+1)].width = widths.get(key, 11)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 42
    ws.freeze_panes = "A3"
    wb.save(out_path)


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTION HYP  (1 ligne par paire STATION HYP ON → OFF)
# ─────────────────────────────────────────────────────────────────────────────

def _extract_hyp(ev, ts_ns_arr, lat_arr, lon_arr, sog_arr, has_nmea):
    """
    1 ligne par paire STATION_HYP ON → OFF.
    Turbidity et Secchi sont attachés à la station active au moment de l'event.
    """
    rows = []

    # Filtrer les events HYP, TURBIDITY, SECCHI
    hyp_ev    = ev[ev["event_type"] == "STATION_HYP"].reset_index(drop=True)
    turbi_ev  = ev[ev["event_type"] == "TURBIDITY"].reset_index(drop=True)
    secchi_ev = ev[ev["event_type"] == "SECCHI"].reset_index(drop=True)

    if hyp_ev.empty:
        return []

    # Construire les sessions HYP : ON → OFF
    sessions = []
    open_row  = None

    for _, row in hyp_ev.iterrows():
        detail = str(row["event_detail"]).strip().upper()
        if "ON" in detail:
            open_row = row
        elif "OFF" in detail and open_row is not None:
            # Extraire le nom de station depuis le detail du ON
            station_name = _station_id_from_detail(str(open_row["event_detail"]), "hyp")
            if not station_name:
                # Fallback sur hyp_station du dataframe
                station_name = str(open_row.get("hyp_station", "")) or "?"

            sessions.append({
                "station_id":   station_name,
                "on_row":       open_row,
                "off_row":      row,
                "on_ts":        pd.Timestamp(_to_ns(open_row["timestamp"])),
                "off_ts":       pd.Timestamp(_to_ns(row["timestamp"])),
                "on_ns":        _to_ns(open_row["timestamp"]),
                "off_ns":       _to_ns(row["timestamp"]),
            })
            open_row = None

    for s in sessions:
        t0, t1 = s["on_ns"], s["off_ns"]
        ts_s, ts_e = s["on_ts"], s["off_ts"]
        dur = (t1 - t0) / 1e9 / 60.0

        if has_nmea:
            lat_s, lon_s, _ = _interp(ts_ns_arr, lat_arr, lon_arr, sog_arr, t0)
            lat_e, lon_e, _ = _interp(ts_ns_arr, lat_arr, lon_arr, sog_arr, t1)
            sog_m            = _mean_sog(ts_ns_arr, sog_arr, t0, t1)
        else:
            lat_s = lon_s = lat_e = lon_e = sog_m = np.nan

        dist  = (_haversine_km(lat_s, lon_s, lat_e, lon_e)
                 if has_nmea and not any(math.isnan(x) for x in [lat_s, lon_s, lat_e, lon_e])
                 else np.nan)
        sog_h = (dist / 1.852 / (dur / 60.0)
                 if not math.isnan(dist) and dur > 0 else np.nan)

        # Turbidity dans la fenêtre de la station
        turbi_in = turbi_ev[
            (turbi_ev["timestamp"] >= ts_s) &
            (turbi_ev["timestamp"] <= ts_e + pd.Timedelta(hours=2))
        ]
        turbi_t1 = turbi_t2 = turbi_t3 = turbi_time = ""
        if not turbi_in.empty:
            tr = turbi_in.iloc[0]
            parsed = _parse_turbidity(str(tr["event_detail"]))
            turbi_t1   = parsed["t1"]
            turbi_t2   = parsed["t2"]
            turbi_t3   = parsed["t3"]
            turbi_time = pd.Timestamp(_to_ns(tr["timestamp"])).strftime("%H:%M:%S")

        # Secchi dans la fenêtre de la station
        secchi_in = secchi_ev[
            (secchi_ev["timestamp"] >= ts_s) &
            (secchi_ev["timestamp"] <= ts_e + pd.Timedelta(hours=2))
        ]
        secchi_val = ""
        if not secchi_in.empty:
            secchi_val = _parse_secchi(str(secchi_in.iloc[0]["event_detail"]))

        rows.append({
            "station_id":    s["station_id"],
            "station_type":  "hyp",
            "date_start":    ts_s.strftime("%Y-%m-%d"),
            "time_start":    ts_s.strftime("%H:%M:%S"),
            "date_end":      ts_e.strftime("%Y-%m-%d"),
            "time_end":      ts_e.strftime("%H:%M:%S"),
            "lat_start_dms": _decimal_to_dms(lat_s),
            "lon_start_dms": _decimal_to_dms(lon_s, True),
            "lat_start_dec": _fmt8(lat_s),
            "lon_start_dec": _fmt8(lon_s),
            "lat_end_dms":   _decimal_to_dms(lat_e),
            "lon_end_dms":   _decimal_to_dms(lon_e, True),
            "lat_end_dec":   _fmt8(lat_e),
            "lon_end_dec":   _fmt8(lon_e),
            "duration_min":  _fmt2(dur),
            "dist_km":       _fmt6(dist),
            "sog_haversine": _fmt4(sog_h),
            "sog_nmea":      _fmt4(sog_m),
            "time_turbi":    turbi_time,
            "turbi_1":       turbi_t1,
            "turbi_2":       turbi_t2,
            "turbi_3":       turbi_t3,
            "secchi":        secchi_val,
        })

    return rows


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTION BIO  (1 ligne par FILTRATION + lignes BUCKET)
# ─────────────────────────────────────────────────────────────────────────────

def _extract_bio(ev, ts_ns_arr, lat_arr, lon_arr, sog_arr, has_nmea):
    """
    - 1 ligne par paire STATION_BIO ON → OFF  (= 1 station)
      Pour chaque station :
        - 1 ligne par paire FILTRATION ON → OFF (avec size, volume, planctospace)
        - 1 ligne par event BUCKET dans la fenêtre de la station
    """
    rows = []

    bio_ev    = ev[ev["event_type"] == "STATION_BIO"].reset_index(drop=True)
    filt_ev   = ev[ev["event_type"] == "FILTRATION"].reset_index(drop=True)
    bucket_ev = ev[ev["event_type"] == "BUCKET"].reset_index(drop=True)

    if bio_ev.empty:
        return []

    # Construire les sessions BIO : ON → OFF
    sessions = []
    open_row  = None

    for _, row in bio_ev.iterrows():
        detail = str(row["event_detail"]).strip().upper()
        if "ON" in detail:
            open_row = row
        elif "OFF" in detail and open_row is not None:
            station_name = _station_id_from_detail(str(open_row["event_detail"]), "bio")
            if not station_name:
                station_name = str(open_row.get("bio_station", "")) or "?"

            sessions.append({
                "station_id": station_name,
                "on_ts":      pd.Timestamp(_to_ns(open_row["timestamp"])),
                "off_ts":     pd.Timestamp(_to_ns(row["timestamp"])),
                "on_ns":      _to_ns(open_row["timestamp"]),
                "off_ns":     _to_ns(row["timestamp"]),
            })
            open_row = None

    filt_counter = {}  # station_id → sub-index for filtrations

    for s in sessions:
        t0_st, t1_st = s["on_ns"], s["off_ns"]
        ts_s_st = s["on_ts"]
        station_name = s["station_id"]

        if station_name not in filt_counter:
            filt_counter[station_name] = 0

        # ── Filtrations dans la fenêtre de la station ──────────────────────
        filts_in = filt_ev[
            (filt_ev["timestamp"] >= s["on_ts"]) &
            (filt_ev["timestamp"] <= s["off_ts"] + pd.Timedelta(hours=1))
        ].reset_index(drop=True)

        filt_open_ns = filt_open_ts = None
        filt_size_on  = ""

        def _add_filtration_row(t0, t1, ts_s, ts_e, filt_size, parsed_off):
            nonlocal rows
            filt_counter[station_name] += 1
            sub = filt_counter[station_name]

            if has_nmea:
                la_s, lo_s, _ = _interp(ts_ns_arr, lat_arr, lon_arr, sog_arr, t0)
                la_e, lo_e, _ = _interp(ts_ns_arr, lat_arr, lon_arr, sog_arr, t1)
                sog_m          = _mean_sog(ts_ns_arr, sog_arr, t0, t1)
            else:
                la_s = lo_s = la_e = lo_e = sog_m = np.nan

            dur  = (t1 - t0) / 1e9 / 60.0
            dist = (_haversine_km(la_s, lo_s, la_e, lo_e)
                    if has_nmea and not any(math.isnan(x) for x in [la_s, lo_s, la_e, lo_e])
                    else np.nan)
            sog_h = (dist / 1.852 / (dur / 60.0)
                     if not math.isnan(dist) and dur > 0 else np.nan)

            size = filt_size or parsed_off.get("size", "")
            s_min, s_max = _size_min_max(size)

            rows.append({
                "station_id":    f"{station_name}_filt{sub}",
                "station_type":  "bio",
                "date_start":    ts_s.strftime("%Y-%m-%d"),
                "time_start":    ts_s.strftime("%H:%M:%S"),
                "date_end":      ts_e.strftime("%Y-%m-%d"),
                "time_end":      ts_e.strftime("%H:%M:%S"),
                "lat_start_dms": _decimal_to_dms(la_s),
                "lon_start_dms": _decimal_to_dms(lo_s, True),
                "lat_start_dec": _fmt8(la_s),
                "lon_start_dec": _fmt8(lo_s),
                "lat_end_dms":   _decimal_to_dms(la_e),
                "lon_end_dms":   _decimal_to_dms(lo_e, True),
                "lat_end_dec":   _fmt8(la_e),
                "lon_end_dec":   _fmt8(lo_e),
                "device":        "Lamprey",
                "mouth_diam":    _MOUTH_DIAM_LAMPREY,
                "depth":         _DEPTH_BIO,
                "size_min":      s_min,
                "size_max":      s_max,
                "duration_min":  _fmt2(dur),
                "dist_km":       _fmt6(dist),
                "sog_haversine": _fmt4(sog_h),
                "sog_nmea":      _fmt4(sog_m),
                # Lamprey analysis
                "vol_lamprey":      size,          # size (micro/nano/pico) — renommé
                "time_filt_start":  ts_s.strftime("%H:%M:%S"),
                "time_filt_end":    ts_e.strftime("%H:%M:%S"),
                "saturation":       parsed_off.get("saturation", ""),
                "vol_lamprey_fil":  parsed_off.get("volume_ml", ""),
                # PlanktoSpace
                "planktospace_st":  parsed_off.get("planctospace", ""),
            })

        for _, frow in filts_in.iterrows():
            detail = str(frow["event_detail"]).strip()
            t_ns   = _to_ns(frow["timestamp"])
            ts     = pd.Timestamp(t_ns)

            if detail.upper().startswith("ON") or "| SIZE" in detail.upper():
                # Peut être " ON | STATION ... | SIZE MICRO"
                if "OFF" not in detail.upper():
                    filt_open_ns  = t_ns
                    filt_open_ts  = ts
                    parsed_on     = _parse_filtration_on(detail)
                    filt_size_on  = parsed_on["size"]

            if "OFF" in detail.upper() and filt_open_ns is not None:
                parsed_off = _parse_filtration_off(detail)
                # Si size pas dans le OFF, utiliser celui du ON
                if not parsed_off["size"]:
                    parsed_off["size"] = filt_size_on
                _add_filtration_row(
                    filt_open_ns, t_ns,
                    filt_open_ts, ts,
                    filt_size_on, parsed_off
                )
                filt_open_ns = filt_open_ts = None
                filt_size_on = ""

        # ── Buckets dans la fenêtre de la station ──────────────────────────
        buckets_in = bucket_ev[
            (bucket_ev["timestamp"] >= s["on_ts"]) &
            (bucket_ev["timestamp"] <= s["off_ts"] + pd.Timedelta(hours=2))
        ].reset_index(drop=True)

        for b_idx, b_row in buckets_in.iterrows():
            b_ns = _to_ns(b_row["timestamp"])
            b_ts = pd.Timestamp(b_ns)
            if has_nmea:
                b_lat, b_lon, b_sog = _interp(ts_ns_arr, lat_arr, lon_arr, sog_arr, b_ns)
            else:
                b_lat = b_lon = b_sog = np.nan

            rows.append({
                "station_id":    f"{station_name}_bucket{b_idx + 1}",
                "station_type":  "bio",
                "date_start":    b_ts.strftime("%Y-%m-%d"),
                "time_start":    b_ts.strftime("%H:%M:%S"),
                "lat_start_dms": _decimal_to_dms(b_lat),
                "lon_start_dms": _decimal_to_dms(b_lon, True),
                "lat_start_dec": _fmt8(b_lat),
                "lon_start_dec": _fmt8(b_lon),
                "device":        f"Bucket {b_idx + 1}",
                "depth":         _DEPTH_BIO,
                "sog_nmea":      _fmt4(b_sog),
                "_row_bg":       _ORANGE,
            })

    return rows


# ─────────────────────────────────────────────────────────────────────────────
# POINT D'ENTRÉE PUBLIC
# ─────────────────────────────────────────────────────────────────────────────

def run_science_export(df, df_events, meta, log_dir: str, base: str):
    departure   = meta.get("departure", "") or "?"
    destination = meta.get("destination", "") or "?"
    leg_label   = f"{departure} → {destination}"
    print(f"  [SCIENCE] Export science — {leg_label}")

    ev = _normalize_events(df_events)
    if ev.empty or "event_type" not in ev.columns:
        print("  [SCIENCE] Aucun événement trouvé.")
        return

    ev = ev.dropna(subset=["timestamp"]).sort_values("timestamp").reset_index(drop=True)

    # Prépare arrays NMEA
    df_nmea = pd.DataFrame()
    if not df.empty and "datetime" in df.columns:
        df_nmea = df.dropna(subset=["datetime"]).copy()
        df_nmea = df_nmea.sort_values("datetime").reset_index(drop=True)
    ts_ns_arr, lat_arr, lon_arr, sog_arr, has_nmea = _build_nmea_arrays(df_nmea)

    hyp_rows = _extract_hyp(ev, ts_ns_arr, lat_arr, lon_arr, sog_arr, has_nmea)
    bio_rows = _extract_bio(ev, ts_ns_arr, lat_arr, lon_arr, sog_arr, has_nmea)

    all_rows = hyp_rows + bio_rows

    if not all_rows:
        print("  [SCIENCE] Aucune station BIO ou HYP détectée.")
        return

    out_path = os.path.join(log_dir, base + "_science.xlsx")
    _write_xlsx(all_rows, out_path, leg_label)

    n_hyp  = len(hyp_rows)
    n_filt = len([r for r in bio_rows if "_filt"   in r.get("station_id","")])
    n_bkt  = len([r for r in bio_rows if "_bucket" in r.get("station_id","")])
    print(f"  [SCIENCE] {n_hyp} station(s) HYP + {n_filt} filtration(s) + {n_bkt} bucket(s) → {out_path}")
    print( "  [SCIENCE] → Vérifier puis copier-coller dans Bio_sampling_Vela_Lab (Google Drive).")
